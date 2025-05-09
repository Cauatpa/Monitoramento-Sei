import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
import time

caminho_arquivo = r"COLOQUE O SEU AQUI"
wb = load_workbook(caminho_arquivo)
ws = wb.active

# Coleta os hyperlinks
links = []
for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
    cell = row[0]
    if cell.hyperlink:
        links.append(cell.hyperlink.target)
    elif isinstance(cell.value, str) and "http" in cell.value:
        links.append(cell.value.strip())
    else:
        links.append(None)

# Remove valores vazios
links = [l for l in links if l]

# Função de extração
def extrair_dados(link):
    headers = {"User-Agent": "Mozilla/5.0"}
    try:
        resp = requests.get(link, headers=headers, timeout=10)
        if resp.status_code != 200:
            return "Erro", "Erro", "Erro", "Erro"

        soup = BeautifulSoup(resp.text, "html.parser")

        numero_processo = soup.find(string=lambda s: s and "SEI-" in s)
        numero_processo = numero_processo.strip() if numero_processo else "Número não encontrado"

        tipo_texto = "Tipo não encontrado"
        data_texto = "Data não encontrada"
        unidade_texto = "Unidade não encontrada"

        for tabela in soup.find_all("table"):
            headers = [th.get_text(strip=True) for th in tabela.find_all("th")]
            if "Tipo" in headers:
                idx_tipo = headers.index("Tipo")
                idx_data = headers.index("Data") if "Data" in headers else 0
                idx_unidade = headers.index("Unidade") if "Unidade" in headers else -1

                linhas = tabela.find_all("tr")[1:]
                if linhas:
                    ultima_linha = linhas[-1]
                    colunas = ultima_linha.find_all("td")
                    if len(colunas) > idx_tipo:
                        tipo_texto = colunas[idx_tipo].get_text(strip=True)
                    if len(colunas) > idx_data:
                        data_texto = colunas[idx_data].get_text(strip=True)
                    if idx_unidade != -1 and len(colunas) > idx_unidade:
                        unidade_texto = colunas[idx_unidade].get_text(strip=True)
                break

        return numero_processo, data_texto, tipo_texto, unidade_texto

    except Exception as e:
        return "Erro", "Erro", "Erro", "Erro"

# Execução
print("\n🔍 Verificando processos SEI:")
for link in links:
    numero, data, tipo, unidade = extrair_dados(link)
    print(f"- {numero} | Última movimentação: {data} | Descrição: {tipo} | Unidade: {unidade}")
    time.sleep(1)
